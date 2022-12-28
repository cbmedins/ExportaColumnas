from tkinter import messagebox, PhotoImage
import openpyxl 
import tkinter.filedialog
import os
import tkinter as tk
import xlrd
import xlsxwriter


# Crea una ventana de tkinter
window = tk.Tk()
window.resizable(0,0)
window.title('Aplicativo para exportar columnas')  
window.iconbitmap('images/icon.ico') 

screen_width = window.winfo_screenwidth()
screen_height = window.winfo_screenheight()
#dimensiones de la ventana
window.geometry('400x600')
# Calcula la posición x e y para centrar la ventana
x_coord = (screen_width/2) - (400/2)
y_coord = (screen_height/2) - (600/2)
# Usa geometry() para mover la ventana a la posición calculada
window.geometry('+%d+%d' % (x_coord, y_coord))


# Crea un widget Frame como separador
separator = tk.Frame(window, height=2, bd=1, relief='sunken')
# Coloca el widget Frame al final de la ventana
separator.pack(side='bottom', fill='x')
# Crea un widget Label para el pie de página
footer_label = tk.Label(window, text='Created by Cristhian Medina')
# Coloca el widget Label debajo del separador
footer_label.pack(side='bottom')



# Crea una función para seleccionar los archivos de Excel
def seleccionar_archivos():
    archivos = tk.filedialog.askopenfilenames(parent=window, title='Seleccionar archivos')

    #Abrir archivo destino
    wb = openpyxl.Workbook()
    # Selecciona la hoja de cálculo activa del nuevo archivo de Excel
    sheet = wb.active
    # Inicializa un contador en 1
    contador = 1

    # Recorre cada archivo de Excel seleccionado en el diálogo
    for file in archivos:
        wb2 = openpyxl.load_workbook(file) #abre el archivo
        sheet2 = wb2['Session 1'] #selecciona la hoja de calculo
        # Recorre cada fila de la hoja de cálculo
        for row in sheet2.rows:
        # Calcula el índice de la columna en la que debes añadir el valor utilizando el contador
            columna = openpyxl.utils.get_column_letter(contador)
        # Añade el valor de la primera columna (que es la columna que quieres copiar) a la fila 1 del archivo de destino en la columna calculada
            sheet[f'{columna}1'] = os.path.basename(file)
            sheet[f'{columna}{row[2].row}'] = row[2].value 
        # Cierra el archivo de Excel
        wb2.close()
        # incrementa el contador a 1
        contador +=1
    # Guarda el archivo de destino
    wb.save('archivo_destino.xlsx')
    messagebox.showinfo(message='Procesado con exito',title='Selecciona los archivos')

def converter():
    messagebox.showinfo(message='Solo se pueden convertir archivos de tipo xsl ya que no son compatibles con la libreria openpyxl',title='Converter')
    # Abrimos el archivo XLS
   # Abrimos el archivo XLS con xlrd
    archivos = tk.filedialog.askopenfilenames(parent=window, title='Seleccionar archivos')
    for file in archivos:
        wb = xlrd.open_workbook(file)

        # Creamos un nuevo archivo XLSX con xlsxwriter
        folder = 'NuevaConversion'
        filename = os.path.basename(file)

        if not os.path.exists(folder):
            os.makedirs(folder)

        wb_xlsx = xlsxwriter.Workbook(folder + '/' + filename + '.xlsx')

        # Copiamos cada hoja del archivo XLS al nuevo archivo XLSX
        for sheet in wb.sheets():
            ws_xlsx = wb_xlsx.add_worksheet(sheet.name)
            for row in range(sheet.nrows):
                for col in range(sheet.ncols):
                    ws_xlsx.write(row, col, sheet.cell_value(row, col))

        # Cerramos el nuevo archivo XLSX
        wb_xlsx.close()
        
    messagebox.showinfo(message='Procesado con exito',title='Convertir archivos')
    
def data_cleaning():
    wb = openpyxl.load_workbook('archivo_destino.xlsx')
    ws = wb['Sheet']
    valores = []

    for col in range(1, ws.max_column + 1):
        subvalores = []
    # Recorrer las filas de la columna
        for row in range(1, ws.max_row + 1):
            # Leer el valor de la celda
            cell_value = ws.cell(row=row, column=col).value
            # Verificar si el valor es mayor que -11
            if cell_value is not None and type(cell_value) is not str and cell_value <= -11.5:
                subvalores.append(cell_value)
        # Si se interrumpió el procesamiento, pasar a la siguiente columna
        else:
            valores.append(subvalores)
            continue
        break
    #print(valores)
 
    wb2 = openpyxl.Workbook()
    sheet = wb2.worksheets[0]
# Recorrer las sublistas de la lista principal
    for i, subvalores in enumerate(valores):
        # Recorrer los valores de la sublista
        for j, valor in enumerate(subvalores):
            # Escribir el valor en la celda correspondiente
            sheet.cell(row=j+1, column=i+1).value = valor

    # Guardar el libro de trabajo
    wb2.save('archivo_limpio.xlsx')
    messagebox.showinfo(message='Procesado con exito',title='Limpieza de datos')

def configuration():
    messagebox.showinfo(message='En mantenimiento',title='Configuración')

def sync():
    wb = openpyxl.load_workbook('archivo_destino.xlsx')
    ws = wb['Sheet']
    fila1 = ws[1]

    wb2 = openpyxl.load_workbook('archivo_limpio.xlsx')
    ws2 = wb2['Sheet']
    
    wb3 = openpyxl.Workbook()
    ws3 = wb3['Sheet']

    for celda in ws[1]:
        ws3[celda.coordinate].value=celda.value

    for i, fila in enumerate(ws2.iter_rows(min_row=1, max_row=ws2.max_row)):
        fila_nueva = i+2
        for celda in fila:
            coordenada=celda.coordinate
            coordenada=coordenada[0]+str(fila_nueva)
            ws3[coordenada].value=celda.value

    wb3.save('archivo_sincronizado.xlsx')

    messagebox.showinfo(message='Procesado con exito',title='Syncronizacion de archivos')

def local():
    # Obtenemos el directorio de la carpeta local del proyecto
    carpeta_local = os.path.dirname(__file__)
    # Abre la carpeta local del proyecto en el explorador de archivos
    os.startfile(carpeta_local)

ResImg1 = PhotoImage(file = 'images/amp82-i72es.png')  
ResImg2 = PhotoImage(file = 'images/aiy8l-mz80y.png')
ResImg3 = PhotoImage(file = 'images/aazfb-yy3p1.png')
ResImg4 = PhotoImage(file = 'images/as8gb-3xzre.png')
ResImg5 = PhotoImage(file = 'images/a3wja-6bdwm.png') 
ResImg6 = PhotoImage(file = 'images/a6g66-ll38e.png') 
#ResImg3 = img3.subsample(24,24)

button1 = tk.Button(text='Selecciona los archivos a exportar',border = 0, image = ResImg1, command=seleccionar_archivos, compound='left') 
button1.place(x=50, y=50)

button2 = tk.Button(text='Convertir archivos',border = 0, image = ResImg2, command=converter, compound='left' )
button2.place(x=50)

button4 = tk.Button(text='Limpieza de datos', border=0, image = ResImg4, command=data_cleaning, compound='left')
button4.place(x=50, y=100)

button3 = tk.Button(text='Configuración',border = 0, image = ResImg3, command=configuration, compound='left' )
button3.place(x=50, y=300)

button5 = tk.Button(text='Sincronizar datos con el archivo destino',border=0,image=ResImg5,command=sync,compound='left')
button5.place(x=50,y=150)

button6 = tk.Button(text='Abrir carpeta del destino',border=0,image=ResImg6,command=local,compound='left')
button6.place(x=50,y=200)


window.mainloop()
